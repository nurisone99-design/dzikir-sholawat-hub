import asyncio
import io
import re
from copy import deepcopy
from types import SimpleNamespace

import pytest
import httpx
from bson import ObjectId
from fastapi import HTTPException
from openpyxl import load_workbook

import server


CABANG_A = ObjectId("64c00000000000000000000a")
CABANG_B = ObjectId("64c00000000000000000000b")
GURU_A = ObjectId("64d00000000000000000000a")
GURU_B = ObjectId("64d00000000000000000000b")
ROLES = (
    "super_admin",
    "penerus_ilmu",
    "ketua_yayasan",
    "viewer_1",
    "admin_cabang",
    "viewer",
    "viewer_2",
)
GLOBAL_ROLES = ("super_admin", "penerus_ilmu", "ketua_yayasan", "viewer_1")
# viewer_2 is branch-scoped for export and cannot export guru at all.
EXPORT_BRANCH_ROLES = ("admin_cabang", "viewer", "viewer_2")


def run(coro):
    return asyncio.run(coro)


def actor(role):
    user = {"id": f"{role}-id", "role": role}
    if role in {"admin_cabang", "viewer", "viewer_2"}:
        user["cabang_id"] = str(CABANG_A)
    return user


def matches(document, query):
    if not query:
        return True
    if "$and" in query and not all(matches(document, item) for item in query["$and"]):
        return False
    if "$or" in query and not any(matches(document, item) for item in query["$or"]):
        return False
    for key, expected in query.items():
        if key in {"$and", "$or"}:
            continue
        actual = document.get(key)
        if isinstance(expected, dict) and "$regex" in expected:
            flags = re.IGNORECASE if "i" in expected.get("$options", "") else 0
            if re.match(expected["$regex"], str(actual or ""), flags) is None:
                return False
        elif isinstance(actual, list):
            if expected not in actual:
                return False
        elif actual != expected:
            return False
    return True


class FakeCursor:
    def __init__(self, documents):
        self.documents = documents

    def sort(self, key=None, direction=1):
        if key:
            self.documents.sort(key=lambda item: item.get(key, ""), reverse=direction < 0)
        return self

    async def to_list(self, limit):
        return deepcopy(self.documents[:limit])


class FakeCollection:
    def __init__(self, documents=()):
        self.documents = deepcopy(list(documents))
        self.find_queries = []
        self.inserted = []

    def find(self, query=None, *_args, **_kwargs):
        query = query or {}
        self.find_queries.append(deepcopy(query))
        return FakeCursor([item for item in self.documents if matches(item, query)])

    async def find_one(self, query, *_args, **_kwargs):
        return next((deepcopy(item) for item in self.documents if matches(item, query)), None)

    async def insert_one(self, document):
        self.inserted.append(deepcopy(document))
        return SimpleNamespace(inserted_id=ObjectId())


class ExportDB:
    def __init__(self):
        self.cabang = FakeCollection([
            {"_id": CABANG_A, "id_cabang": "CABANG_A", "kota": "Cabang A", "guru_id": str(GURU_A)},
            {"_id": CABANG_B, "id_cabang": "CABANG_B", "kota": "Cabang B", "guru_id": str(GURU_B)},
        ])
        self.jamaah = FakeCollection([
            {"_id": ObjectId(), "nama": "JAMAAH_A_MARKER", "gender": "Laki-laki", "nama_orang_tua": "ORTU_A", "ijazah_nama_dalam": ["Al Yakin", "As Sirrul Jannah", "Abdi Sholihin"], "cabang_id": str(CABANG_A)},
            {"_id": ObjectId(), "nama": "JAMAAH_B_MARKER", "gender": "Perempuan", "nama_orang_tua": "ORTU_B", "ijazah_nama_dalam": ["Nama B"], "cabang_id": str(CABANG_B)},
        ])
        self.guru = FakeCollection([
            {"_id": GURU_A, "id_guru": "GURU_A", "nama": "GURU_A_MARKER", "cabang_id": str(CABANG_A)},
            {"_id": GURU_B, "id_guru": "GURU_B", "nama": "GURU_B_MARKER", "cabang_ids": [str(CABANG_B)]},
        ])
        self.pengurus = FakeCollection([
            {"_id": ObjectId(), "id_pengurus": "PA", "nama": "PENGURUS_A_MARKER", "cabang_id": str(CABANG_A)},
            {"_id": ObjectId(), "id_pengurus": "PB", "nama": "PENGURUS_B_MARKER", "cabang_id": str(CABANG_B)},
        ])
        self.agenda = FakeCollection([
            {"_id": ObjectId(), "judul": "AGENDA_A_MARKER", "cabang_id": str(CABANG_A)},
            {"_id": ObjectId(), "judul": "AGENDA_B_MARKER", "cabang_id": str(CABANG_B)},
        ])
        self.galeri = FakeCollection([
            {"_id": ObjectId(), "judul": "GALERI_A_MARKER", "cabang_id": str(CABANG_A)},
            {"_id": ObjectId(), "judul": "GALERI_B_MARKER", "cabang_id": str(CABANG_B)},
        ])
        self.pengumuman = FakeCollection([])
        self.settings = FakeCollection([{"_id": ObjectId(), "key": "yayasan", "nama": "Yayasan Test"}])
        self.audit_logs = FakeCollection([])

    def __getitem__(self, name):
        return getattr(self, name)


@pytest.fixture
def export_db(monkeypatch):
    database = ExportDB()
    monkeypatch.setattr(server, "db", database)
    return database


async def response_bytes(response):
    chunks = []
    async for chunk in response.body_iterator:
        chunks.append(chunk)
    return b"".join(chunks)


def make_export(entity, user, **overrides):
    params = {
        "format": "xlsx", "cabang": None, "cabang_id": None,
        "gender": None, "columns": None, "fields": None, "preset": None,
        "user": user,
    }
    params.update(overrides)
    return run(server.export_data(entity, **params))


def worksheet_values(response):
    content = run(response_bytes(response))
    sheet = load_workbook(io.BytesIO(content), read_only=True).active
    return [list(row) for row in sheet.iter_rows(values_only=True)]


@pytest.mark.parametrize("entity,marker_a,marker_b", [
    ("jamaah", "JAMAAH_A_MARKER", "JAMAAH_B_MARKER"),
    ("guru", "GURU_A_MARKER", "GURU_B_MARKER"),
    ("pengurus", "PENGURUS_A_MARKER", "PENGURUS_B_MARKER"),
])
@pytest.mark.parametrize("role", [r for r in ROLES if r != "viewer_2"])
def test_branch_owned_export_scope(export_db, entity, marker_a, marker_b, role):
    response = make_export(entity, actor(role), fields="nama")
    values = worksheet_values(response)
    flattened = {cell for row in values for cell in row if cell}

    assert marker_a in flattened
    assert (marker_b in flattened) is (role in GLOBAL_ROLES)


@pytest.mark.parametrize("entity,marker_a,marker_b", [
    ("jamaah", "JAMAAH_A_MARKER", "JAMAAH_B_MARKER"),
    ("pengurus", "PENGURUS_A_MARKER", "PENGURUS_B_MARKER"),
])
@pytest.mark.parametrize("role", EXPORT_BRANCH_ROLES)
def test_branch_roles_export_only_their_branch(export_db, entity, marker_a, marker_b, role):
    response = make_export(entity, actor(role), fields="nama")
    flattened = {cell for row in worksheet_values(response) for cell in row if cell}

    assert marker_a in flattened
    assert marker_b not in flattened


@pytest.mark.parametrize("entity", ["guru", "agenda", "galeri", "pengumuman"])
def test_viewer2_cannot_export_non_whitelisted_entity(export_db, entity):
    with pytest.raises(HTTPException) as exc:
        make_export(entity, actor("viewer_2"), fields="nama")
    assert exc.value.status_code == 403


@pytest.mark.parametrize("entity,field", [
    ("jamaah", "nama"),
    ("cabang", "kota"),
    ("pengurus", "nama"),
])
def test_viewer2_can_export_whitelisted_entity(export_db, entity, field):
    response = make_export(entity, actor("viewer_2"), fields=field)
    assert response.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.mark.parametrize("entity,marker_a,marker_b", [
    ("agenda", "AGENDA_A_MARKER", "AGENDA_B_MARKER"),
    ("galeri", "GALERI_A_MARKER", "GALERI_B_MARKER"),
])
@pytest.mark.parametrize("role", [r for r in ROLES if r != "viewer_2"])
def test_global_read_export_scope(export_db, entity, marker_a, marker_b, role):
    response = make_export(entity, actor(role), fields="judul")
    flattened = {cell for row in worksheet_values(response) for cell in row if cell}
    assert {marker_a, marker_b} <= flattened


@pytest.mark.parametrize("entity", ["jamaah", "guru", "pengurus"])
@pytest.mark.parametrize("parameter", ["cabang", "cabang_id"])
def test_client_branch_filter_cannot_expand_scope(export_db, entity, parameter):
    response = make_export(entity, actor("admin_cabang"), fields="nama", **{parameter: str(CABANG_B)})
    flattened = {cell for row in worksheet_values(response) for cell in row if cell}
    assert any(str(value).endswith("A_MARKER") for value in flattened)
    assert not any(str(value).endswith("B_MARKER") for value in flattened)


@pytest.mark.parametrize("entity", ["agenda", "galeri"])
def test_branch_filter_does_not_reduce_global_read_for_branch_role(export_db, entity):
    response = make_export(entity, actor("admin_cabang"), cabang_id=str(CABANG_B), fields="judul")
    markers = {cell for row in worksheet_values(response) for cell in row if cell and "MARKER" in str(cell)}
    assert len(markers) == 2


def test_custom_columns_are_whitelisted_and_keep_requested_order(export_db):
    rows = worksheet_values(make_export("jamaah", actor("super_admin"), fields="gender,nama"))
    assert rows[0] == ["Gender", "Nama Lengkap"]
    assert all(len(row) == 2 for row in rows)


def test_illegal_sensitive_column_is_rejected(export_db):
    with pytest.raises(HTTPException) as exc:
        make_export("jamaah", actor("super_admin"), fields='["nama", "password_hash"]')
    assert exc.value.status_code == 400


def test_metadata_contains_safe_presets_and_exact_usulan_definition():
    metadata = run(server.get_export_fields("jamaah", actor("viewer")))
    presets = {item["key"]: item for item in metadata["presets"]}
    assert {"default", "data_dasar", "data_keanggotaan", "usulan_nama_dalam"} <= set(presets)
    assert presets["usulan_nama_dalam"] == {
        "key": "usulan_nama_dalam",
        "label": "Usulan Nama Dalam",
        "fields": ["gender", "nama", "nama_orang_tua", "ijazah_nama_dalam"],
    }
    assert "password_hash" not in {item["key"] for item in metadata["fields"]}


def test_usulan_xlsx_title_headers_columns_and_single_cell_list(export_db):
    response = make_export(
        "jamaah", actor("admin_cabang"), preset="usulan_nama_dalam"
    )
    rows = worksheet_values(response)
    assert rows[0][0] == f"USULAN NAMA DALAM ({server.datetime.now(server.timezone.utc).year})"
    assert rows[2][:2] == ["Cabang", "Cabang A"]
    assert rows[3][:2] == ["Guru Pembimbing", "GURU_A_MARKER"]
    assert rows[5] == ["Gender", "Nama", "Nama Orang Tua", "Ijazah Nama Dalam"]
    assert rows[6][3] == "Al Yakin, As Sirrul Jannah, Abdi Sholihin"
    assert len(rows) == 7


def test_usulan_xlsx_multi_branch_and_teacher_headers(export_db):
    rows = worksheet_values(make_export("jamaah", actor("super_admin"), preset="usulan_nama_dalam"))
    assert rows[2][:2] == ["Cabang", "Semua Cabang"]
    assert rows[3][:2] == ["Guru Pembimbing", "Semua Guru Pembimbing"]


def test_usulan_preset_allows_custom_field_override(export_db):
    rows = worksheet_values(make_export(
        "jamaah", actor("admin_cabang"), preset="usulan_nama_dalam", fields="nama,gender"
    ))
    assert rows[5] == ["Nama", "Gender"]


def test_usulan_pdf_content_is_extractable(export_db):
    PdfReader = pytest.importorskip("pypdf").PdfReader
    response = make_export("jamaah", actor("admin_cabang"), format="pdf", preset="usulan_nama_dalam")
    content = run(response_bytes(response))
    text = "\n".join(page.extract_text() or "" for page in PdfReader(io.BytesIO(content)).pages)
    assert f"USULAN NAMA DALAM ({server.datetime.now(server.timezone.utc).year})" in text
    assert "Cabang : Cabang A" in text
    assert "Guru Pembimbing : GURU_A_MARKER" in text
    assert "Nama Orang Tua" in text
    assert "JAMAAH_A_MARKER" in text
    assert "Al Yakin, As Sirrul Jannah, Abdi Sholihin" in text


def test_usulan_pdf_is_generated(export_db):
    response = make_export("jamaah", actor("admin_cabang"), format="pdf", preset="usulan_nama_dalam")
    content = run(response_bytes(response))
    assert response.media_type == "application/pdf"
    assert content.startswith(b"%PDF-")
    assert len(content) > 1000


@pytest.mark.parametrize("parameter", ["cabang_id", "branch", "branch_id"])
def test_http_query_manipulation_cannot_expand_jamaah_scope(export_db, parameter):
    async def current_user_override():
        return actor("admin_cabang")

    async def request_export():
        server.app.dependency_overrides[server.get_current_user] = current_user_override
        try:
            transport = httpx.ASGITransport(app=server.app)
            async with httpx.AsyncClient(transport=transport, base_url="http://localhost") as client:
                return await client.get(
                    f"/api/export/jamaah?format=xlsx&fields=nama&{parameter}={CABANG_B}"
                )
        finally:
            server.app.dependency_overrides.pop(server.get_current_user, None)

    response = run(request_export())
    assert response.status_code == 200
    sheet = load_workbook(io.BytesIO(response.content), read_only=True).active
    values = {cell for row in sheet.iter_rows(values_only=True) for cell in row if cell}
    assert "JAMAAH_A_MARKER" in values
    assert "JAMAAH_B_MARKER" not in values


def test_existing_export_without_fields_or_preset_uses_old_defaults(export_db):
    rows = worksheet_values(make_export("jamaah", actor("super_admin")))
    assert rows[0] == [server.COLUMN_TITLE_MAP[key] for key in server.DEFAULT_COLUMNS["jamaah"]]
    assert len(rows[0]) == len(server.DEFAULT_COLUMNS["jamaah"])


@pytest.mark.parametrize("role", ["penerus_ilmu", "ketua_yayasan", "viewer_1", "viewer", "viewer_2"])
def test_export_does_not_grant_write_permission(role):
    assert server.require_write
    with pytest.raises(HTTPException) as exc:
        server.require_write(actor(role))
    assert exc.value.status_code == 403
